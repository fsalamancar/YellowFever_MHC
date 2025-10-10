from pathlib import Path
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def color_heatmap_excel(csv_path: str,
                        excel_path: str | None = None,
                        colored_path: str | None = None,
                        index_col: int = 0) -> str:
    """
    Convierte un CSV a .xlsx, aplica color a las celdas con emojis 🟢/🔵 y elimina los emojis.
    Devuelve la ruta al fichero coloreado.
    """
    from pathlib import Path
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    csv_p = Path(csv_path)
    if excel_path is None:
        excel_path = str(csv_p.with_suffix(".xlsx"))
    if colored_path is None:
        colored_path = str(Path(excel_path).with_name(Path(excel_path).stem + "_coloreado.xlsx"))

    # Leer CSV y guardar como Excel
    df = pd.read_csv(csv_p, index_col=index_col)
    # Limpiar nombres de columnas: quitar la secuencia literal ".*" y espacios
    df.columns = [re.sub(r"\.\d+$", "", str(c)).strip() for c in df.columns]
    df.to_excel(excel_path, index=True)

    # Abrir y colorear según emojis
    wb = load_workbook(excel_path)
    ws = wb.active

    fill_green = PatternFill(start_color="2BB800", end_color="2BB800", fill_type="solid")
    fill_blue  = PatternFill(start_color="06A1C4", end_color="06A1C4", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2):  # ignora encabezado e índice
        for cell in row:
            if cell.value == "🟢":
                cell.fill = fill_green
            elif cell.value == "🔵":
                cell.fill = fill_blue
            # Eliminar emoji del valor
            if cell.value in ["🟢", "🔵"]:
                cell.value = ""

    wb.save(colored_path)
    print("Saved colored Excel:", colored_path)
    return colored_path

if __name__ == "__main__":
    from pathlib import Path

    csv_input = input("Ruta al CSV (o deja vacío para cancelar): ").strip()
    if not csv_input:
        print("Operación cancelada.")
    else:
        csv_p = Path(csv_input)
        while not csv_p.exists():
            print("CSV no encontrado:", csv_p)
            csv_input = input("Ingresa una ruta válida al CSV (o deja vacío para salir): ").strip()
            if not csv_input:
                print("Operación cancelada.")
                raise SystemExit(1)
            csv_p = Path(csv_input)

        excel_input = input("Ruta de salida .xlsx (Enter para usar la misma carpeta): ").strip() or None
        colored_input = input("Ruta de salida .xlsx coloreado (Enter para usar <nombre>_coloreado.xlsx): ").strip() or None

        result = color_heatmap_excel(str(csv_p), excel_input, colored_input)
        print("Hecho:", result)